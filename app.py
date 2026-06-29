from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
import shutil
from datetime import datetime
import json
from functools import wraps
import hashlib
import hmac

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'change-this-secret-key-in-production')
app.config['SESSION_PERMANENT'] = False
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'

EXCEL_DIR = 'excel_sheets'
MONTHS_FILE = 'months_data.json'
BACKUP_DIR = 'backups'
DELETED_MONTHS_DIR = os.path.join(BACKUP_DIR, 'deleted_months')
BACKUP_MONTHS_FILE = os.path.join(BACKUP_DIR, 'backup_months.json')

def hash_password(password):
    salt = os.getenv('PASSWORD_SALT', 'uta-default-salt-change-me')
    return hashlib.sha256((salt + str(password)).encode('utf-8')).hexdigest()

DEFAULT_USERS = {
    'thanzeel': hash_password(os.getenv('THANZEEL_PASSWORD', 'utams2179')),
    'kaiff': hash_password(os.getenv('KAIFF_PASSWORD', 'utams2179')),
}

def load_users():
    raw = os.getenv('UTA_USERS_JSON')
    if not raw:
        return DEFAULT_USERS
    try:
        plain_users = json.loads(raw)
        return {u.lower(): hash_password(p) for u, p in plain_users.items()}
    except Exception:
        return DEFAULT_USERS

USERS = load_users()

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if not session.get('user'):
            if request.path.startswith('/get_') or request.method != 'GET':
                return jsonify({'success': False, 'message': 'Unauthorized. Please login again.'}), 401
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return wrapper

def create_backup(reason, month_name=None):
    os.makedirs(BACKUP_DIR, exist_ok=True)
    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_folder = os.path.join(BACKUP_DIR, f'{stamp}_{reason}')
    os.makedirs(backup_folder, exist_ok=True)
    if os.path.exists(MONTHS_FILE):
        shutil.copy2(MONTHS_FILE, os.path.join(backup_folder, 'months_data.json'))
    if month_name:
        excel_file = os.path.join(EXCEL_DIR, f'{month_name}.xlsx')
        if os.path.exists(excel_file):
            shutil.copy2(excel_file, os.path.join(backup_folder, f'{month_name}.xlsx'))
    return backup_folder

# Create Excel directory if it doesn't exist
if not os.path.exists(EXCEL_DIR):
    os.makedirs(EXCEL_DIR)
if not os.path.exists(BACKUP_DIR):
    os.makedirs(BACKUP_DIR)
if not os.path.exists(DELETED_MONTHS_DIR):
    os.makedirs(DELETED_MONTHS_DIR)



def load_backup_months():
    if os.path.exists(BACKUP_MONTHS_FILE):
        with open(BACKUP_MONTHS_FILE, 'r') as f:
            return json.load(f)
    return {}

def save_backup_months(data):
    with open(BACKUP_MONTHS_FILE, 'w') as f:
        json.dump(data, f, indent=2)

def move_month_to_backup(month_name):
    """Move a deleted month into the Backup page instead of permanently deleting it."""
    months_data = load_months_data()
    if month_name not in months_data:
        return False
    os.makedirs(DELETED_MONTHS_DIR, exist_ok=True)
    month_data = months_data[month_name].copy()
    source_file = os.path.join(EXCEL_DIR, f"{month_name}.xlsx")
    backup_file = os.path.join(DELETED_MONTHS_DIR, f"{month_name}.xlsx")
    if os.path.exists(source_file):
        shutil.move(source_file, backup_file)
    month_data['deleted_date'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    month_data['backup_filename'] = backup_file
    month_data['status'] = 'backup'
    backup_months = load_backup_months()
    backup_months[month_name] = month_data
    save_backup_months(backup_months)
    del months_data[month_name]
    save_months_data(months_data)
    return True

def load_months_data():
    """Load months data from JSON file"""
    if os.path.exists(MONTHS_FILE):
        with open(MONTHS_FILE, 'r') as f:
            return json.load(f)
    return {}

def save_months_data(data):
    """Save months data to JSON file"""
    with open(MONTHS_FILE, 'w') as f:
        json.dump(data, f, indent=2)

def parse_amount(value, default=0):
    """Parse amounts entered with commas, e.g. 10,000."""
    try:
        return float(str(value).replace(",", "").strip() or default)
    except Exception:
        return float(default)

def format_amount(value):
    """Format amounts with comma grouping across the app."""
    try:
        number = float(value or 0)
    except (TypeError, ValueError):
        number = 0
    if number.is_integer():
        return f"{int(number):,}"
    return f"{number:,.2f}"

app.jinja_env.filters["amount"] = format_amount

def apply_excel_amount_format(filename):
    """Remove decimal .00 display from amount columns before saving or downloading."""
    if not os.path.exists(filename):
        return
    wb = load_workbook(filename)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        for col in [6, 7, 9]:
            cell = row[col - 1]
            cell.number_format = "#,##0"
            if isinstance(cell.value, float) and cell.value.is_integer():
                cell.value = int(cell.value)
    wb.save(filename)
    wb.close()


def get_current_month():
    """Get current month-year string"""
    return datetime.now().strftime("%B_%Y")

def get_current_month_display():
    """Get current month display name"""
    return datetime.now().strftime("%B %Y")

def create_monthly_sheet(month_name, opening_balance, opening_date=None):
    """Create a new Excel sheet for a month"""
    filename = os.path.join(EXCEL_DIR, f"{month_name}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = "Records"
    
    # Add headers
    headers = ['No', 'Ref No', 'Date', 'Subject', 'Pass No', 'In Payment', 'Out Payment', 'Sub Agent', 'Balance']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="1a1a2e", end_color="1a1a2e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Add opening balance row
    if not opening_date:
        opening_date = datetime.now().strftime("%Y-%m-%d")
    ws.append([1, 'OPENING', opening_date, 'OPENING BALANCE', '', 0, 0, '', opening_balance])
    
    # Style opening balance row
    opening_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    opening_font = Font(bold=True, color="2E7D32")
    for col in range(1, 10):
        cell = ws.cell(2, col)
        cell.fill = opening_fill
        cell.font = opening_font
    
    for col in [6, 7, 9]:
        ws.cell(2, col).number_format = '#,##0'

    # Adjust column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 15
    ws.column_dimensions['I'].width = 15
    
    wb.save(filename)
    return filename

def close_monthly_sheet(month_name):
    """Close a monthly sheet (mark as completed)"""
    months_data = load_months_data()
    if month_name in months_data:
        months_data[month_name]['status'] = 'closed'
        months_data[month_name]['closed_date'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        save_months_data(months_data)

def delete_month_sheet(month_name):
    """Delete a monthly sheet completely"""
    try:
        months_data = load_months_data()
        
        # Delete Excel file
        filename = os.path.join(EXCEL_DIR, f"{month_name}.xlsx")
        if os.path.exists(filename):
            os.remove(filename)
        
        # Remove from months data
        if month_name in months_data:
            del months_data[month_name]
            save_months_data(months_data)
        
        return True
    except Exception as e:
        print(f"Error deleting month: {e}")
        return False

def is_month_active(month_name):
    """Check if a month sheet is active"""
    months_data = load_months_data()
    if month_name in months_data:
        return months_data[month_name].get('status') == 'active'
    return False

def get_active_month():
    """Get the currently active month"""
    months_data = load_months_data()
    for month, data in months_data.items():
        if data.get('status') == 'active':
            return month
    return None

def get_next_reference_number(month_name):
    """Get the next reference number for a specific month"""
    filename = os.path.join(EXCEL_DIR, f"{month_name}.xlsx")
    if not os.path.exists(filename):
        return 1
    
    wb = load_workbook(filename)
    ws = wb.active
    max_num = 0
    
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[1] and isinstance(row[1], str) and row[1].startswith('UTA-'):
            try:
                num = int(row[1].split('-')[1])
                max_num = max(max_num, num)
            except:
                continue
    
    wb.close()
    return max_num + 1

def get_next_row_number(month_name):
    """Get the next row number"""
    filename = os.path.join(EXCEL_DIR, f"{month_name}.xlsx")
    if not os.path.exists(filename):
        return 2
    
    wb = load_workbook(filename)
    ws = wb.active
    max_no = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] and isinstance(row[0], (int, float)) and row[0] != 'OPENING':
            max_no = max(max_no, int(row[0]))
    
    wb.close()
    return max_no + 1

def get_current_balance(month_name):
    """Get the current balance from the last row"""
    filename = os.path.join(EXCEL_DIR, f"{month_name}.xlsx")
    if not os.path.exists(filename):
        return 0
    
    wb = load_workbook(filename)
    ws = wb.active
    last_balance = 0
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[8] is not None and row[1] != 'OPENING':
            last_balance = float(row[8]) if row[8] else 0
        elif row[1] == 'OPENING':
            last_balance = float(row[8]) if row[8] else 0
    
    wb.close()
    return last_balance

@app.route('/')
def index():
    if not session.get('user'):
        return redirect(url_for('login_page'))
    return redirect(url_for('dashboard'))

@app.route('/login', methods=['GET'])
def login_page():
    if session.get('user'):
        return redirect(url_for('dashboard'))
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def login():
    data = request.get_json(silent=True) or {}
    username = str(data.get('username', '')).strip().lower()
    password = str(data.get('password', ''))
    stored_hash = USERS.get(username)
    if stored_hash and hmac.compare_digest(stored_hash, hash_password(password)):
        session.clear()
        session.permanent = False
        session['user'] = username
        session['login_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        return jsonify({'success': True, 'message': 'Login successful'})
    return jsonify({'success': False, 'message': 'Invalid username or password'}), 401

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))

@app.route('/dashboard')
@login_required
def dashboard():
    """Dashboard showing all months"""
    months_data = load_months_data()
    active_month = get_active_month()
    current_month_display = get_current_month_display()
    return render_template('dashboard.html', 
                         months_data=months_data, 
                         active_month=active_month,
                         current_month_display=current_month_display,
                         current_user=session.get('user'))

@app.route('/create_month', methods=['POST'])
@login_required
def create_month():
    """Create a new month sheet"""
    try:
        data = request.json
        opening_balance = parse_amount(data.get('opening_balance', 0))
        creation_date_str = data.get('creation_date', '').strip()

        # Validate and parse the creation date
        if creation_date_str:
            try:
                creation_date = datetime.strptime(creation_date_str, "%Y-%m-%d")
            except ValueError:
                return jsonify({'success': False, 'message': 'Invalid date format. Please use YYYY-MM-DD.'})
        else:
            creation_date = datetime.now()

        # Derive month key and display name from the chosen date
        current_month = creation_date.strftime("%B_%Y")
        month_display = creation_date.strftime("%B %Y")

        # Check if month already exists
        months_data = load_months_data()
        if current_month in months_data:
            return jsonify({'success': False, 'message': f'Month {month_display} already exists!'})

        # Create Excel file using the chosen date as the opening balance date
        opening_date_str = creation_date.strftime("%Y-%m-%d")
        filename = create_monthly_sheet(current_month, opening_balance, opening_date=opening_date_str)

        # Save month data — store the chosen date (not server now()) as created_date
        months_data[current_month] = {
            'display_name': month_display,
            'created_date': creation_date.strftime("%Y-%m-%d") + " 00:00:00",
            'opening_balance': opening_balance,
            'status': 'active',
            'filename': filename
        }
        save_months_data(months_data)

        session['active_month'] = current_month

        return jsonify({
            'success': True,
            'message': f'Month {month_display} created successfully!',
            'month_name': current_month
        })

    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/delete_month', methods=['POST'])
@login_required
def delete_month():
    """Delete a month sheet completely"""
    try:
        data = request.json
        month_name = data.get('month_name')
        
        if not month_name:
            return jsonify({'success': False, 'message': 'Month name not provided'})
        
        # Check if it's the active month
        active_month = get_active_month()
        if active_month == month_name:
            return jsonify({'success': False, 'message': 'Cannot delete active month! Please end the month first.'})
        
        # Move the month to the Backup page instead of deleting permanently
        create_backup('delete_month', month_name)
        if move_month_to_backup(month_name):
            return jsonify({'success': True, 'message': 'Month moved to Backup successfully!'})
        else:
            return jsonify({'success': False, 'message': 'Error moving month to Backup'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/end_month', methods=['POST'])
@login_required
def end_month():
    """Close the current month"""
    try:
        active_month = get_active_month()
        if not active_month:
            return jsonify({'success': False, 'message': 'No active month found!'})
        
        close_monthly_sheet(active_month)
        session.pop('active_month', None)
        
        return jsonify({'success': True, 'message': f'Month closed successfully!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/enter')
@login_required
def enter_records():
    """Enter records page"""
    active_month = get_active_month()
    if not active_month:
        return render_template('no_active_month.html')
    months_data = load_months_data()
    return render_template('enter_records.html', 
                         active_month=active_month, 
                         month_display=months_data[active_month]['display_name'])

@app.route('/view')
@login_required
def view_records():
    """View records page"""
    months_data = load_months_data()
    return render_template('view_records.html', months_data=months_data)

@app.route('/view_month/<month_name>')
@login_required
def view_month_records(month_name):
    """View specific month records"""
    months_data = load_months_data()
    if month_name not in months_data:
        return "Month not found", 404
    return render_template('view_month.html', 
                         month_name=month_name, 
                         month_data=months_data[month_name])

@app.route('/delete_record_from_month', methods=['POST'])
@login_required
def delete_record_from_month():
    """Delete a single record from a specific month"""
    try:
        data = request.json
        month_name = data.get('month_name')
        ref_no = data.get('ref_no') or str(data.get('record_no', ''))

        if not month_name or not ref_no:
            return jsonify({'success': False, 'message': 'Missing required data'})
        
        filename = os.path.join(EXCEL_DIR, f"{month_name}.xlsx")
        if not os.path.exists(filename):
            return jsonify({'success': False, 'message': 'Month file not found'})
        
        wb = load_workbook(filename)
        ws = wb.active
        row_to_delete = None
        for row in range(3, ws.max_row + 1):
            if str(ws.cell(row, 2).value or '').strip() == str(ref_no).strip():
                row_to_delete = row
                break
        
        if row_to_delete:
            create_backup('delete_record', month_name)
            ws.delete_rows(row_to_delete)

            # Re-sort by date, reassign Ref Nos sequentially, and recalculate balances
            recalculate_and_sort_sheet(ws)

            wb.save(filename)
            wb.close()
            return jsonify({'success': True, 'message': f'Record deleted and references re-indexed successfully!'})
        else:
            wb.close()
            return jsonify({'success': False, 'message': 'Record not found'})
        
    except Exception as e:
        print(f"Error deleting record: {e}")
        return jsonify({'success': False, 'message': f'Error deleting: {str(e)}'})

@app.route('/update_record', methods=['POST'])
@login_required
def update_record():
    """Update a single record in a specific month"""
    try:
        data = request.json
        month_name = data.get('month_name')
        ref_no = data.get('ref_no')

        if not month_name or not ref_no:
            return jsonify({'success': False, 'message': 'Missing required data'})

        filename = os.path.join(EXCEL_DIR, f"{month_name}.xlsx")
        if not os.path.exists(filename):
            return jsonify({'success': False, 'message': 'Month file not found'})

        wb = load_workbook(filename)
        ws = wb.active
        row_to_update = None
        for row in range(3, ws.max_row + 1):
            if str(ws.cell(row, 2).value or '').strip() == str(ref_no).strip():
                row_to_update = row
                break

        if row_to_update is None:
            wb.close()
            return jsonify({'success': False, 'message': 'Record not found'})

        create_backup('update_record', month_name)

        # Update the fields (keep ref_no and row number unchanged for now)
        ws.cell(row_to_update, 3).value = data.get('date', ws.cell(row_to_update, 3).value)
        ws.cell(row_to_update, 4).value = data.get('subject', ws.cell(row_to_update, 4).value)
        ws.cell(row_to_update, 5).value = data.get('pass_no', ws.cell(row_to_update, 5).value)
        ws.cell(row_to_update, 6).value = parse_amount(data.get('in_payment', 0))
        ws.cell(row_to_update, 7).value = parse_amount(data.get('out_payment', 0))
        ws.cell(row_to_update, 8).value = data.get('sub_agent', ws.cell(row_to_update, 8).value)

        # Recalculate all balances and re-sort
        recalculate_and_sort_sheet(ws)

        wb.save(filename)
        wb.close()
        return jsonify({'success': True, 'message': 'Record updated successfully!'})

    except Exception as e:
        print(f"Error updating record: {e}")
        return jsonify({'success': False, 'message': f'Error updating: {str(e)}'})


@app.route('/get_month_records/<month_name>')
@login_required
def get_month_records(month_name):
    """Get records for a specific month"""
    try:
        filename = os.path.join(EXCEL_DIR, f"{month_name}.xlsx")
        if not os.path.exists(filename):
            return jsonify({'records': []})
        
        wb = load_workbook(filename)
        ws = wb.active
        
        records = []
        for row in ws.iter_rows(min_row=3, values_only=True):
            if row[0] and row[0] != 'OPENING' and row[0] is not None:
                records.append({
                    'no': row[0],
                    'ref_no': row[1] if row[1] else '',
                    'date': str(row[2]) if row[2] else '',
                    'subject': row[3] if row[3] else '',
                    'pass_no': row[4] if row[4] else '',
                    'in_payment': float(row[5]) if row[5] else 0,
                    'out_payment': float(row[6]) if row[6] else 0,
                    'sub_agent': row[7] if row[7] else '',
                    'balance': float(row[8]) if row[8] else 0
                })
        
        wb.close()
        return jsonify({'records': records})
        
    except Exception as e:
        print(f"Error getting records: {e}")
        return jsonify({'records': [], 'error': str(e)})

def recalculate_and_sort_sheet(ws):
    """Sort all transaction rows by date (opening row stays first), renumber, and recalculate balances."""
    # Collect opening row (row 2, col B == 'OPENING')
    opening_row_data = None
    transaction_rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        row = list(row)
        if row[1] == 'OPENING':
            opening_row_data = row
        elif row[0] is not None:
            transaction_rows.append(row)

    if opening_row_data is None:
        return

    def parse_date_safe(row):
        try:
            date_val = row[2]
            if isinstance(date_val, str):
                return datetime.strptime(date_val, '%Y-%m-%d')
            elif hasattr(date_val, 'year'):
                return datetime(date_val.year, date_val.month, date_val.day)
        except Exception:
            pass
        return datetime(9999, 12, 31)

    transaction_rows.sort(key=parse_date_safe)

    # Clear rows from row 2 onward, then re-write
    ws.delete_rows(2, ws.max_row)

    # Write opening row back
    ws.append(opening_row_data)
    opening_row_idx = ws.max_row
    opening_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    opening_font = Font(bold=True, color="2E7D32")
    for col in range(1, 10):
        cell = ws.cell(opening_row_idx, col)
        cell.fill = opening_fill
        cell.font = opening_font
    for col in [6, 7, 9]:
        ws.cell(opening_row_idx, col).number_format = '#,##0'

    # Write sorted transactions with recalculated balances and reassigned ref numbers
    balance = float(opening_row_data[8]) if opening_row_data[8] is not None else 0
    for i, row in enumerate(transaction_rows):
        in_payment = float(row[5]) if row[5] else 0
        out_payment = float(row[6]) if row[6] else 0
        balance = balance + in_payment - out_payment

        # Reassign ref number based on sorted position (1-based)
        new_ref_no = f"UTA-{str(i + 1).zfill(2)}"

        new_row = [
            i + 2,         # Row No (#)
            new_ref_no,    # Ref No — reassigned sequentially after date sort
            row[2],        # Date
            row[3],        # Subject
            row[4],        # Pass No
            in_payment,
            out_payment,
            row[7],        # Sub Agent
            balance
        ]
        ws.append(new_row)
        row_idx = ws.max_row
        for col in [6, 7, 9]:
            ws.cell(row_idx, col).number_format = '#,##0'


@app.route('/save_records', methods=['POST'])
@login_required
def save_records():
    """Save records to active month sheet"""
    try:
        active_month = get_active_month()
        if not active_month:
            return jsonify({'success': False, 'message': 'No active month! Please create a new month first.'})
        
        data = request.json
        records = data.get('records', [])
        
        if not records:
            return jsonify({'success': False, 'message': 'No records to save'})
        
        filename = os.path.join(EXCEL_DIR, f"{active_month}.xlsx")
        if not os.path.exists(filename):
            return jsonify({'success': False, 'message': 'Month file not found!'})
        
        wb = load_workbook(filename)
        ws = wb.active
        
        next_ref_num = get_next_reference_number(active_month)
        current_balance = get_current_balance(active_month)
        
        saved_count = 0
        ref_offset = 0
        for record in records:
            if not record.get('date') or not record.get('subject'):
                continue
            
            ref_no = f"UTA-{str(next_ref_num + ref_offset).zfill(2)}"
            in_payment = parse_amount(record.get('in_payment', 0))
            out_payment = parse_amount(record.get('out_payment', 0))
            new_balance = current_balance + in_payment - out_payment

            # Row number is a placeholder; recalculate_and_sort_sheet will fix it
            row = [
                0,
                ref_no,
                record.get('date', ''),
                record.get('subject', ''),
                record.get('pass_no', ''),
                in_payment,
                out_payment,
                record.get('sub_agent', ''),
                new_balance
            ]
            ws.append(row)
            current_balance = new_balance
            saved_count += 1
            ref_offset += 1

        # Sort all rows by date and recalculate balances in one pass
        recalculate_and_sort_sheet(ws)

        wb.save(filename)
        wb.close()
        
        return jsonify({
            'success': True,
            'message': f'Successfully saved {saved_count} records!',
            'next_reference': f"UTA-{str(next_ref_num + saved_count).zfill(2)}"
        })
        
    except Exception as e:
        print(f"Error saving records: {e}")
        return jsonify({'success': False, 'message': f'Error saving: {str(e)}'})
@app.route('/get_next_reference')
@login_required
def get_next_reference():
    """Get the next reference number for display"""
    active_month = get_active_month()
    if not active_month:
        return jsonify({'next_reference': 'No Active Month'})
    
    next_num = get_next_reference_number(active_month)
    return jsonify({'next_reference': f"UTA-{str(next_num).zfill(2)}"})

@app.route('/get_stats')
@login_required
def get_stats():
    """Get statistics for dashboard — shows active month data only."""
    months_data = load_months_data()
    total_months = len(months_data)
    active_month = get_active_month()

    total_records = 0
    total_in_payment = 0
    total_out_payment = 0
    total_references = 0
    net_balance = 0

    if active_month:
        filename = os.path.join(EXCEL_DIR, f"{active_month}.xlsx")
        if os.path.exists(filename):
            wb = load_workbook(filename)
            ws = wb.active
            opening_balance = 0

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[1] == 'OPENING':
                    opening_balance = float(row[8]) if row[8] else 0
                elif row[0] is not None:
                    total_records += 1
                    total_in_payment += float(row[5]) if row[5] else 0
                    total_out_payment += float(row[6]) if row[6] else 0
                    if row[1] and row[1] != 'OPENING':
                        total_references += 1

            wb.close()
            net_balance = opening_balance + total_in_payment - total_out_payment

    return jsonify({
        'total_months': total_months,
        'total_records': total_records,
        'total_in_payment': total_in_payment,
        'total_out_payment': total_out_payment,
        'total_references': total_references,
        'net_balance': net_balance,
        'active_month': active_month,
        'active_month_display': months_data.get(active_month, {}).get('display_name', 'None') if active_month else 'None'
    })

@app.route('/get_month_stats/<month_name>')
@login_required
def get_month_stats(month_name):
    """Get statistics for a specific month"""
    try:
        filename = os.path.join(EXCEL_DIR, f"{month_name}.xlsx")
        if not os.path.exists(filename):
            return jsonify({'error': 'Month not found'}), 404
        
        wb = load_workbook(filename)
        ws = wb.active
        
        total_records = 0
        total_in_payment = 0
        total_out_payment = 0
        opening_balance = 0
        closing_balance = 0
        total_references = set()
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == 'OPENING':
                opening_balance = float(row[8]) if row[8] else 0
            elif row[0] and row[0] != 'OPENING' and row[0] is not None:
                total_records += 1
                total_in_payment += float(row[5]) if row[5] else 0
                total_out_payment += float(row[6]) if row[6] else 0
                if row[1]:
                    total_references.add(row[1])
                closing_balance = float(row[8]) if row[8] else 0
        
        wb.close()
        
        return jsonify({
            'total_records': total_records,
            'total_in_payment': total_in_payment,
            'total_out_payment': total_out_payment,
            'total_references': len(total_references),
            'opening_balance': opening_balance,
            'closing_balance': closing_balance,
            'net_change': closing_balance - opening_balance
        })
        
    except Exception as e:
        print(f"Error getting month stats: {e}")
        return jsonify({'error': str(e)}), 500


@app.route('/backup')
@login_required
def backup_page():
    backup_months = load_backup_months()
    return render_template('backup.html', backup_months=backup_months)


@app.route('/restore_backup_month', methods=['POST'])
@login_required
def restore_backup_month():
    try:
        data = request.json
        month_name = data.get('month_name')
        if not month_name:
            return jsonify({'success': False, 'message': 'Month name not provided'})

        backup_months = load_backup_months()
        if month_name not in backup_months:
            return jsonify({'success': False, 'message': 'Backup month not found'})

        months_data = load_months_data()
        if month_name in months_data:
            return jsonify({'success': False, 'message': 'This month already exists in records'})

        os.makedirs(EXCEL_DIR, exist_ok=True)
        backup_data = backup_months[month_name].copy()
        backup_file = backup_data.get('backup_filename') or os.path.join(DELETED_MONTHS_DIR, f"{month_name}.xlsx")
        restore_file = os.path.join(EXCEL_DIR, f"{month_name}.xlsx")

        if os.path.exists(backup_file):
            shutil.move(backup_file, restore_file)

        backup_data.pop('deleted_date', None)
        backup_data.pop('backup_filename', None)
        backup_data['status'] = 'closed'
        backup_data['filename'] = restore_file

        months_data[month_name] = backup_data
        save_months_data(months_data)

        del backup_months[month_name]
        save_backup_months(backup_months)

        return jsonify({'success': True, 'message': 'Month restored from Backup successfully!'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/delete_backup_month', methods=['POST'])
@login_required
def delete_backup_month():
    try:
        data = request.json
        month_name = data.get('month_name')
        if not month_name:
            return jsonify({'success': False, 'message': 'Month name not provided'})
        backup_months = load_backup_months()
        if month_name not in backup_months:
            return jsonify({'success': False, 'message': 'Backup month not found'})
        backup_file = backup_months[month_name].get('backup_filename') or os.path.join(DELETED_MONTHS_DIR, f"{month_name}.xlsx")
        if os.path.exists(backup_file):
            os.remove(backup_file)
        del backup_months[month_name]
        save_backup_months(backup_months)
        return jsonify({'success': True, 'message': 'Backup month deleted permanently!'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@app.route('/download_excel/<month_name>')
@login_required
def download_excel(month_name):
    """Download a specific month's Excel file"""
    try:
        filename = os.path.join(EXCEL_DIR, f"{month_name}.xlsx")
        if os.path.exists(filename):
            apply_excel_amount_format(filename)
            return send_file(
                filename, 
                as_attachment=True, 
                download_name=f'UTA_{month_name}.xlsx'
            )
        else:
            return jsonify({'error': 'File not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    print("\n" + "="*60)
    print("📊 UTA Manpower Service - Application Started")
    print("="*60)
    print(f"📁 Excel Sheets Directory: {EXCEL_DIR}")
    print(f"🌐 Server: http://localhost:5000")
    print("\n✅ Month-based sheet management enabled")
    print("✅ Automatic balance calculation")
    print("✅ Delete month and record functionality enabled")
    print("="*60 + "\n")
    
    # Try port 5000, if busy use 5001
    port = 5000
    try:
        app.run(debug=True, port=port, host='127.0.0.1')
    except OSError as e:
        if "Address already in use" in str(e):
            print(f"⚠️ Port {port} is busy. Trying port 5001...")
            port = 5001
            app.run(debug=True, port=port, host='127.0.0.1')
        else:
            raise